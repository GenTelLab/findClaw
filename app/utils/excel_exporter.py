import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_LIST = [
    "IP 地址", "端口", "智能体类型", "版本", "置信度",
    "置信分数", "匹配关键词", "匹配规则", "发现来源", "命中证据", "发现时间", "是否新增",
]

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

HEADER_FILL = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


class ExcelExporter:

    @staticmethod
    def export(result_list: list[dict]) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "扫描结果"

        _write_header(ws)
        _write_data(ws, result_list)
        _auto_width(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


def _write_header(ws) -> None:
    for col_idx, header in enumerate(HEADER_LIST, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN


def _write_data(ws, result_list: list[dict]) -> None:
    for row_idx, item in enumerate(result_list, 2):
        ws.cell(row=row_idx, column=1, value=item.get("ip", ""))
        ws.cell(row=row_idx, column=2, value=item.get("port", ""))
        ws.cell(row=row_idx, column=3, value=item.get("claw_type", ""))
        ws.cell(row=row_idx, column=4, value=item.get("claw_version", ""))
        ws.cell(row=row_idx, column=5, value=item.get("confidence", ""))
        ws.cell(row=row_idx, column=6, value=item.get("confidence_score", 0))
        ws.cell(row=row_idx, column=7, value=item.get("matched_keyword", ""))
        ws.cell(row=row_idx, column=8, value=item.get("matched_rule", ""))
        ws.cell(row=row_idx, column=9, value=item.get("discovery_source", ""))
        ws.cell(row=row_idx, column=10, value=item.get("evidence", ""))

        discovered = item.get("discovered_at")
        if isinstance(discovered, datetime):
            ws.cell(row=row_idx, column=11, value=discovered.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            ws.cell(row=row_idx, column=11, value=str(discovered or ""))

        ws.cell(row=row_idx, column=12, value="是" if item.get("is_new") else "否")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
